"""
Find answers in 六级词汇_题目到答案.xlsx by matching partial or full question text.

Usage:
    python match_answers.py "部分或完整的题目文本"
    python match_answers.py "text" -n 5           # top 5 matches
    python match_answers.py "text" -f other.xlsx  # custom file
"""

from __future__ import annotations

import argparse
import difflib
import re
from pathlib import Path
from typing import Dict, Iterable, List, Tuple

from openpyxl import load_workbook

try:
    import wordninja

    _split_words = wordninja.split
except Exception:  # optional dependency
    _split_words = None


def normalize(text: str) -> str:
    """Lowercase and strip non-alphanumeric characters for fuzzy matching."""
    return re.sub(r"[^a-z0-9]+", "", str(text).lower())


def locate_workbook(explicit: str | None) -> Path:
    if explicit:
        path = Path(explicit).expanduser()
        if not path.exists():
            raise FileNotFoundError(f"Workbook not found: {path}")
        return path

    folder = Path(__file__).resolve().parent
    # Prefer the known filename but fall back to the first .xlsx in the folder.
    preferred = folder / "六级词汇_题目到答案.xlsx"
    if preferred.exists():
        return preferred

    for candidate in folder.glob("*.xlsx"):
        return candidate

    raise FileNotFoundError("No .xlsx workbook found in script directory.")


def load_rows(path: Path) -> List[Dict[str, str]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = [str(h).strip() if h else "" for h in header_row]

    rows: List[Dict[str, str]] = []
    for values in ws.iter_rows(min_row=2, values_only=True):
        row = {h: v for h, v in zip(headers, values)}
        if not row.get("题目文本"):
            continue
        rows.append(row)
    wb.close()
    return rows


def score_matches(query: str, rows: Iterable[Dict[str, str]], top_n: int) -> List[Tuple[float, Dict[str, str]]]:
    norm_query = normalize(query)
    matches: List[Tuple[float, Dict[str, str]]] = []

    for row in rows:
        text = row.get("题目文本(分词)") or row.get("题目文本", "")
        norm_text = normalize(text)
        if not norm_text:
            continue

        # Exact/substring match gets perfect score; else use similarity ratio.
        if norm_query and norm_query in norm_text:
            score = 1.0
        else:
            score = difflib.SequenceMatcher(None, norm_query, norm_text).ratio()
        matches.append((score, row))

    matches.sort(key=lambda item: item[0], reverse=True)
    return matches[:top_n]


def format_match(score: float, row: Dict[str, str]) -> str:
    answer_letter = row.get("答案字母", "")
    answer_word = row.get("答案单词", "")
    question_id = row.get("题号", "")
    question_text = row.get("题目文本", "")
    full_question = row.get("完整题目+选项", "")
    segmented = row.get("题目文本(分词)") or segment_text(question_text)

    lines = [
        f"[score {score:.3f}] 题号: {question_id}  答案: {answer_letter} ({answer_word})",
        f"题目文本: {question_text}",
    ]
    if segmented and segmented != question_text:
        lines.append(f"分词后: {segmented}")
    if full_question:
        lines.append(f"完整题目+选项: {full_question}")
    return "\n".join(lines)


def segment_text(text: str) -> str:
    """Best-effort word segmentation for display."""
    if not text:
        return ""
    cleaned = re.sub(r"[^A-Za-z0-9]+", " ", text).strip()
    if not cleaned:
        return text
    # If wordninja is available, use it per token; otherwise return cleaned.
    if _split_words:
        tokens: List[str] = []
        for part in cleaned.split():
            tokens.extend(_split_words(part))
        return " ".join(tokens)
    return cleaned


def print_matches(query: str, rows: List[Dict[str, str]], top_n: int) -> None:
    matches = score_matches(query, rows, top_n=top_n)
    if not matches:
        print("未找到匹配结果。")
        return
    print(f"查询: {query}")
    print("-" * 60)
    for score, row in matches:
        print(format_match(score, row))
        print("-" * 60)


def main() -> None:
    parser = argparse.ArgumentParser(description="Match CET-6 vocab questions to answers.")
    parser.add_argument("query", nargs="?", help="部分或全部题目文本，用于匹配答案（留空则进入交互模式）")
    parser.add_argument("-f", "--file", dest="file", help="Excel 文件路径（默认同目录下的六级词汇_题目到答案.xlsx）")
    parser.add_argument("-n", "--top", dest="top", type=int, default=3, help="展示匹配度最高的前 N 条结果 (默认 3)")
    args = parser.parse_args()

    workbook = locate_workbook(args.file)
    rows = load_rows(workbook)

    # If no query provided, start a prompt loop.
    if not args.query:
        print(f"使用文件: {workbook}")
        print("输入部分或完整题目文本后回车进行查询，空行退出。")
        while True:
            try:
                query = input("query> ").strip()
            except (EOFError, KeyboardInterrupt):
                print()
                break
            if not query:
                break
            print_matches(query, rows, top_n=args.top)
        return

    print(f"使用文件: {workbook}")
    print_matches(args.query, rows, top_n=args.top)


if __name__ == "__main__":
    main()
